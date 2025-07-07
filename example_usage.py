#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
KBreaker for Excel - Example Usage
=================================

This file demonstrates various ways to use the KBreaker tool
for unlocking Excel files safely and professionally.
"""

import os
import sys
from pathlib import Path
from kbreaker import KBreaker

def example_basic_usage():
    """Basic usage example"""
    print("🔓 Example 1: Basic Usage")
    print("-" * 30)
    
    input_file = "protected_workbook.xlsx"
    
    # Check if example file exists
    if not os.path.exists(input_file):
        print(f"⚠️ Example file '{input_file}' not found")
        print("   Create a password-protected Excel file for testing")
        return
    
    # Process the file
    with KBreaker() as kb:
        success, message = kb.process_file(input_file)
        
        if success:
            print(f"✅ {message}")
        else:
            print(f"❌ {message}")

def example_custom_output():
    """Example with custom output path"""
    print("\n🔓 Example 2: Custom Output Path")
    print("-" * 35)
    
    input_file = "protected_workbook.xlsx"
    output_file = "my_unlocked_file.xlsx"
    
    if not os.path.exists(input_file):
        print(f"⚠️ Example file '{input_file}' not found")
        return
    
    with KBreaker() as kb:
        success, message = kb.process_file(input_file, output_file)
        
        if success:
            print(f"✅ {message}")
        else:
            print(f"❌ {message}")

def example_batch_processing():
    """Example of processing multiple files"""
    print("\n🔓 Example 3: Batch Processing")
    print("-" * 32)
    
    # List of files to process
    input_files = [
        "file1.xlsx",
        "file2.xlsx", 
        "file3.xlsx"
    ]
    
    results = []
    
    with KBreaker() as kb:
        for input_file in input_files:
            if os.path.exists(input_file):
                print(f"\n📁 Processing: {input_file}")
                success, message = kb.process_file(input_file)
                results.append((input_file, success, message))
                
                if success:
                    print(f"✅ Success: {input_file}")
                else:
                    print(f"❌ Failed: {input_file} - {message}")
            else:
                print(f"⚠️ File not found: {input_file}")
                results.append((input_file, False, "File not found"))
    
    # Summary
    print("\n📊 Batch Processing Summary:")
    print("-" * 30)
    successful = sum(1 for _, success, _ in results if success)
    total = len(results)
    print(f"✅ Successful: {successful}/{total}")
    print(f"❌ Failed: {total - successful}/{total}")

def example_error_handling():
    """Example with comprehensive error handling"""
    print("\n🔓 Example 4: Error Handling")
    print("-" * 30)
    
    input_file = "nonexistent_file.xlsx"
    
    try:
        with KBreaker() as kb:
            success, message = kb.process_file(input_file)
            
            if success:
                print(f"✅ Success: {message}")
            else:
                print(f"❌ Error: {message}")
                
                # Handle specific error types
                if "not found" in message.lower():
                    print("💡 Tip: Check if the file path is correct")
                elif "vba" in message.lower():
                    print("💡 Tip: Enable VBA project access in Excel")
                elif "excel" in message.lower():
                    print("💡 Tip: Ensure Excel is properly installed")
                    
    except Exception as e:
        print(f"❌ Unexpected error: {str(e)}")
        print("💡 Check the log file for detailed error information")

def example_file_validation():
    """Example with file validation"""
    print("\n🔓 Example 5: File Validation")
    print("-" * 32)
    
    test_files = [
        "document.xlsx",    # Valid Excel file
        "document.txt",     # Invalid file type
        "nonexistent.xlsx", # File doesn't exist
    ]
    
    for file_path in test_files:
        print(f"\n📁 Validating: {file_path}")
        
        # Check if file exists
        if not os.path.exists(file_path):
            print(f"❌ File not found: {file_path}")
            continue
        
        # Check file extension
        if not file_path.lower().endswith(('.xlsx', '.xls')):
            print(f"❌ Invalid file type: {file_path}")
            print("   Only .xlsx and .xls files are supported")
            continue
        
        # Check file size
        file_size = os.path.getsize(file_path)
        if file_size == 0:
            print(f"❌ Empty file: {file_path}")
            continue
        
        print(f"✅ Valid Excel file: {file_path} ({file_size} bytes)")
        
        # Process the file
        with KBreaker() as kb:
            success, message = kb.process_file(file_path)
            if success:
                print(f"✅ Unlocked successfully")
            else:
                print(f"❌ Failed to unlock: {message}")

def create_test_file():
    """Helper function to create a test Excel file"""
    print("\n🔧 Creating Test File")
    print("-" * 20)
    
    try:
        import openpyxl
        from openpyxl.workbook.protection import WorkbookProtection
        
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Protected Sheet"
        
        # Add some data
        ws['A1'] = "This is a protected Excel file"
        ws['A2'] = "Created for KBreaker testing"
        ws['B1'] = "Column B"
        ws['B2'] = "More data"
        
        # Protect the sheet (this is just for demonstration)
        # Note: openpyxl protection is different from Excel's built-in protection
        ws.protection.sheet = True
        ws.protection.password = "test123"
        
        # Save the file
        test_file = "test_protected.xlsx"
        wb.save(test_file)
        
        print(f"✅ Test file created: {test_file}")
        print("💡 Note: For real testing, manually protect the file in Excel")
        
    except ImportError:
        print("⚠️ openpyxl not installed. Install with: pip install openpyxl")
    except Exception as e:
        print(f"❌ Failed to create test file: {str(e)}")

def main():
    """Run all examples"""
    print("🔓 KBreaker for Excel - Usage Examples")
    print("=" * 45)
    
    # Run examples
    example_basic_usage()
    example_custom_output()
    example_batch_processing()
    example_error_handling()
    example_file_validation()
    
    # Offer to create test file
    print("\n" + "=" * 45)
    response = input("Create a test Excel file? (y/n): ").lower().strip()
    if response in ['y', 'yes']:
        create_test_file()
    
    print("\n🎉 Examples completed!")
    print("💡 Check kbreaker.log for detailed operation logs")

if __name__ == "__main__":
    main()